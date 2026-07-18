"""
nwn_gui_frontend.py
---------------------
Tkinter frontend for the NWN combat log analyzer. All parsing logic
lives in nwn_log_logic.py - this file only builds the window and
displays results.
"""

import os
import re
import glob
import ctypes
import platform
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl import Workbook

from nwn_log_logic import (
    analyze_nwn_log, load_config, save_config,
    DEFAULT_CHARACTERS, IGNORED_NAMES,
)

LEGEND = {
    "Attack Record":
        "Swings = attacks this entity made while attacking "
        "(H = Hits landed, M = Misses). Taken = attacks this "
        "entity received while being attacked (H = Hits landed "
        "against it, M = Misses against it).",
    "Save":
        "Bonus = save bonus range added to the d20 roll. "
        "DC = Difficulty Class the roll must meet or beat. "
        "Needs X+ = the minimum raw d20 result (before adding "
        "bonus) needed to beat the toughest DC seen, using the "
        "lowest bonus seen in the log - it is neither DC nor "
        "bonus alone, but DC minus bonus. "
        "S = successful saves, F = failed saves.",
    "Damage":
        "Dealt = damage of this type this entity inflicted. "
        "Taken = damage of this type this entity received.",
    "Damage Type":
        "Dealt = total damage of this type this monster "
        "inflicted across all its attacks.",
    "Mitigation":
        "DR = Damage Resistance, a flat amount of damage "
        "absorbed per hit. Immune = damage fully negated by "
        "immunity. Concealed = number of attacks that missed "
        "due to concealment (e.g. invisibility, fog).",
    "Threat Rolls":
        "Count = number of threat (aggro) rolls logged. "
        "Max = highest threat roll total seen. "
        "Avg = average threat roll total.",
    "Cast":
        "Number of times this entity cast this spell during "
        "the session.",
    "Per-Attack Damage":
        "Min/Max/Count = smallest, largest, and total number "
        "of individual damage instances dealt by this monster.",
    "Kills":
        "Number of kills this entity scored, and the name(s) "
        "of the victim(s).",
}


def _fix_windows_dpi_scaling():
    if platform.system() != "Windows":
        return
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass


class NWNAnalyzerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("NWN Combat Log Analyzer")
        self.geometry("1150x720")
        self.cfg = load_config()
        self.selected_files = []
        self.stats_pc = {}
        self.stats_m = {}
        self.extra_stats = {}
        self.characters = []

        self._build_top_bar()
        self._build_char_box()
        self._build_tabs()

    # ---------- Top bar ----------
    def _build_top_bar(self):
        frame = ttk.Frame(self)
        frame.pack(fill="x", padx=10, pady=8)

        ttk.Button(
            frame, text="Select Log File(s)",
            command=self.pick_files
        ).pack(side="left")
        ttk.Button(
            frame, text="Use Default Folder",
            command=self.use_default_folder
        ).pack(side="left", padx=6)
        ttk.Button(
            frame, text="Set Default Folder",
            command=self.set_default_folder
        ).pack(side="left", padx=6)

        self.files_label = ttk.Label(
            frame, text="No files selected", foreground="gray")
        self.files_label.pack(side="left", padx=10)

        ttk.Button(
            frame, text="Analyze", command=self.run_analysis
        ).pack(side="right")
        ttk.Button(
            frame, text="Export Excel", command=self.export_excel
        ).pack(side="right", padx=6)
        ttk.Button(
            frame, text="View Events", command=self.open_events_window
        ).pack(side="right")

    # ---------- Character roster ----------
    def _build_char_box(self):
        frame = ttk.LabelFrame(self, text="Player / Toon Roster")
        frame.pack(fill="x", padx=10, pady=4)

        list_frame = ttk.Frame(frame)
        list_frame.pack(
            side="left", fill="both", expand=True, padx=6, pady=6
        )

        self.char_listbox = tk.Listbox(
            list_frame, height=5, selectmode="extended"
        )
        self.char_listbox.pack(side="left", fill="both", expand=True)
        vsb = ttk.Scrollbar(
            list_frame, orient="vertical",
            command=self.char_listbox.yview)
        self.char_listbox.configure(yscrollcommand=vsb.set)
        vsb.pack(side="left", fill="y")

        for name in self.cfg.get("characters", DEFAULT_CHARACTERS):
            self.char_listbox.insert("end", name)

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(side="left", fill="y", padx=6, pady=6)

        self.char_entry = ttk.Entry(btn_frame, width=22)
        self.char_entry.pack(pady=(0, 4))
        self.char_entry.bind("<Return>", lambda e: self.add_character())

        ttk.Button(
            btn_frame, text="Add",
            command=self.add_character).pack(fill="x", pady=2)
        ttk.Button(
            btn_frame, text="Remove Selected",
            command=self.remove_character).pack(fill="x", pady=2)
        ttk.Button(
            btn_frame, text="Save as Default Roster",
            command=self.save_default_roster).pack(
                fill="x", pady=(10, 2))
        ttk.Button(
            btn_frame, text="Reset to Built-in Default",
            command=self.reset_default_roster).pack(fill="x", pady=2)

    def add_character(self):
        name = self.char_entry.get().strip()
        if name:
            self.char_listbox.insert("end", name)
            self.char_entry.delete(0, "end")

    def remove_character(self):
        for idx in reversed(self.char_listbox.curselection()):
            self.char_listbox.delete(idx)

    def get_characters(self):
        return list(self.char_listbox.get(0, "end"))

    def save_default_roster(self):
        characters = self.get_characters()
        if not characters:
            messagebox.showwarning(
                "Empty roster", "Add at least one character first.")
            return
        self.cfg["characters"] = characters
        save_config(self.cfg)
        messagebox.showinfo(
            "Saved",
            "Default roster updated ({} character(s)).".format(
                len(characters)))

    def reset_default_roster(self):
        self.char_listbox.delete(0, "end")
        for name in DEFAULT_CHARACTERS:
            self.char_listbox.insert("end", name)

    # ---------- File selection ----------
    def pick_files(self):
        paths = filedialog.askopenfilenames(
            title="Select NWN log file(s)",
            initialdir=self.cfg.get("default_log_dir", os.getcwd()),
            filetypes=[("Log files", "*.txt *.log"), ("All files", "*.*")]
        )
        if paths:
            self.selected_files = list(paths)
            self.files_label.config(
                text="{} file(s) selected".format(len(paths)))

    def use_default_folder(self):
        folder = self.cfg.get("default_log_dir", os.getcwd())
        found = sorted(
            glob.glob(os.path.join(folder, "nwclientLog*.txt")),
            key=lambda p: int(
                re.search(r"(\d+)", os.path.basename(p)).group(1)
            ) if re.search(r"(\d+)", os.path.basename(p)) else 0
        )
        if not found:
            messagebox.showwarning(
                "No files found",
                "No 'nwclientLog*.txt' files found in:\n" + folder)
            return
        self.selected_files = found
        self.files_label.config(
            text="{} file(s) from default folder".format(len(found)))

    def set_default_folder(self):
        folder = filedialog.askdirectory(title="Select default log folder")
        if folder:
            self.cfg["default_log_dir"] = folder
            save_config(self.cfg)
            messagebox.showinfo(
                "Saved", "Default log folder updated to:\n" + folder)

    # ---------- Tabs ----------
    def _build_tabs(self):
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=6)

        self.tree_char = self._make_tree_tab(
            "Character Stats",
            ["Kills", "Total Dmg", "Avg/Hit", "Detail"])
        self.tree_mon = self._make_tree_tab(
            "Monster Stats",
            ["Max AB", "AC Range", "Avg Dmg", "Detail"])
        self.tree_saves = self._make_flat_tab(
            "Saving Throws (All)",
            ["Name", "Category", "Check", "Count",
             "Save Bonus Range", "DC Range",
             "Min Roll Needed", "Success/Fail"])
        self.tree_mit = self._make_flat_tab(
            "Mitigation & Threat (All)",
            ["Name", "Category", "DR", "Immunity", "Concealed",
             "Avg Conceal %", "Threat Count", "Threat Max", "Threat Avg"])
        self.tree_dmg_types = self._make_flat_tab(
            "Damage Types (All)",
            ["Name", "Category", "Damage Type",
             "Dealt", "Taken"])
        self.tree_leaderboard = self._make_flat_tab(
            "Session Leaderboard",
            ["Rank", "Character", "Total Damage",
             "Kills", "Avg/Hit"])

    def _make_tree_tab(self, title, columns):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text=title)

        tree_frame = ttk.Frame(tab)
        tree_frame.pack(fill="both", expand=True)

        tree = ttk.Treeview(
            tree_frame, columns=columns, show="tree headings")
        tree.heading("#0", text="Name / Detail")
        tree.column("#0", width=220, anchor="w")
        for c in columns:
            tree.heading(c, text=c)
            tree.column(c, width=110, anchor="center")
        vsb = ttk.Scrollbar(
            tree_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        info_var = tk.StringVar(value="Select a row to see details here.")
        info_bar = ttk.Label(
            tab, textvariable=info_var, anchor="w",
            wraplength=1100, padding=(6, 4), relief="sunken")
        info_bar.pack(fill="x", side="bottom")

        tree.bind(
            "<<TreeviewSelect>>",
            lambda e, t=tree, v=info_var: self._on_tree_select(t, v))
        return tree

    def _on_tree_select(self, tree, info_var):
        sel = tree.selection()
        if not sel:
            return
        label = tree.item(sel[0], "text").strip()
        key = label.split(":")[0].strip()
        desc = LEGEND.get(key)
        if desc:
            info_var.set("{}: {}".format(key, desc))
        else:
            info_var.set("")

    def _make_flat_tab(self, title, columns):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text=title)
        tree = ttk.Treeview(tab, columns=columns, show="headings")
        for c in columns:
            tree.heading(c, text=c)
            tree.column(c, width=120, anchor="center")
        vsb = ttk.Scrollbar(tab, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        return tree

    # ---------- Analysis ----------
    def run_analysis(self):
        if not self.selected_files:
            messagebox.showwarning(
                "No files",
                "Please select log file(s) or use the default folder.")
            return
        characters = self.get_characters()
        if not characters:
            messagebox.showwarning(
                "No characters", "Please add at least one character name.")
            return

        (files_processed, errors, stats_pc, stats_m,
         extra_stats) = analyze_nwn_log(self.selected_files, characters)

        if errors:
            messagebox.showwarning(
                "Missing files", "Could not find:\n" + "\n".join(errors))
        if files_processed == 0:
            messagebox.showerror("Error", "No log files could be processed.")
            return

        self.stats_pc = stats_pc
        self.stats_m = stats_m
        self.extra_stats = extra_stats
        self.characters = characters

        self._fill_character_tab(stats_pc, characters, extra_stats)
        self._fill_monster_tab(stats_m, characters, extra_stats)
        self._fill_saves_tab(extra_stats, characters)
        self._fill_mitigation_tab(extra_stats, characters)
        self._fill_damage_types_tab(extra_stats, characters)
        self._fill_leaderboard_tab(stats_pc)
        messagebox.showinfo(
            "Done", "Processed {} file(s) successfully.".format(
                files_processed))

    def _clear_tree(self, tree):
        for row in tree.get_children():
            tree.delete(row)

    def _fill_character_tab(self, stats_pc, characters, extra_stats):
        self._clear_tree(self.tree_char)
        for p in characters:
            d = stats_pc[p]
            avg = d["dmg_tot"] / d["hits_count"] if d["hits_count"] else 0
            parent = self.tree_char.insert(
                "", "end", text=p, values=(
                    d["kills"], d["dmg_tot"], round(avg, 2), ""))

            ex = extra_stats.get(p)
            if not ex:
                continue

            dealt = ex["dmg_dealt_types"]
            taken = ex["dmg_taken_types"]
            all_types = sorted(set(dealt.keys()) | set(taken.keys()))
            for dtype in all_types:
                self.tree_char.insert(
                    parent, "end",
                    text="  Damage: " + dtype,
                    values=("", "", "",
                            "Dealt {} / Taken {}".format(
                                dealt.get(dtype, 0),
                                taken.get(dtype, 0))))

            for check_key, res in sorted(ex["checks"].items()):
                totals = res["totals"]
                if not totals:
                    continue
                dcs = res["dcs"]
                bonuses = res["bonuses"]
                min_roll_needed = max(max(dcs) - min(bonuses), 1)
                self.tree_char.insert(
                    parent, "end",
                    text="  Save: " + check_key,
                    values=("", "", "",
                            "Bonus {}-{} / DC {}-{} / "
                            "Needs {}+ / {}S-{}F".format(
                                min(bonuses), max(bonuses),
                                min(dcs), max(dcs), min_roll_needed,
                                res["success"], res["fail"])))

            if ex["dr_absorbed"] or ex["immunity_absorbed"] or \
                    ex["concealed_against"]:
                self.tree_char.insert(
                    parent, "end",
                    text="  Mitigation",
                    values=("", "", "",
                            "DR {} / Immune {} / Concealed {}x".format(
                                ex["dr_absorbed"], ex["immunity_absorbed"],
                                ex["concealed_against"])))

            if ex["threat_rolls"]:
                tr = ex["threat_rolls"]
                self.tree_char.insert(
                    parent, "end",
                    text="  Threat Rolls",
                    values=("", "", "",
                            "Count {} / Max {} / Avg {:.2f}".format(
                                len(tr), max(tr), sum(tr) / len(tr))))

            spells = ex.get("spells_cast")
            if spells:
                for spell, count in sorted(spells.items()):
                    self.tree_char.insert(
                        parent, "end",
                        text="  Cast: " + spell,
                        values=("", "", "", "{}x".format(count)))

            atk_h = ex.get("atk_hits", 0)
            atk_m = ex.get("atk_misses", 0)
            def_h = ex.get("def_hits", 0)
            def_m = ex.get("def_misses", 0)
            if atk_h or atk_m or def_h or def_m:
                self.tree_char.insert(
                    parent, "end",
                    text="  Attack Record",
                    values=("", "", "",
                            "Swings {}H-{}M / Taken {}H-{}M".format(
                                atk_h, atk_m, def_h, def_m)))


                    
    def _fill_monster_tab(self, stats_m, characters, extra_stats):

        self._clear_tree(self.tree_mon)
        for m, d in sorted(stats_m.items()):
            if m in characters or m in IGNORED_NAMES:
                continue
            max_ab = max(d["ab"]) if d["ab"] else 0
            hi_miss = max(d["misses_ac"]) + 1 if d["misses_ac"] else "?"
            lo_hit = min(d["hits_ac"]) if d["hits_ac"] else "?"
            avg_m = d["dmg_val"] / d["dmg_count"] if d["dmg_count"] else 0

            parent = self.tree_mon.insert(
                "", "end", text=m, values=(
                    max_ab, "{} - {}".format(hi_miss, lo_hit),
                    round(avg_m, 2), ""))

            hits = d.get("dmg_hits", [])
            if hits:
                self.tree_mon.insert(
                    parent, "end",
                    text="  Per-Attack Damage",
                    values=("", "", "",
                            "Min {} / Max {} / Count {}".format(
                                min(hits), max(hits), len(hits))))

            ex = extra_stats.get(m)
            if ex:
                for check_key, res in sorted(ex["checks"].items()):
                    totals = res["totals"]
                    if not totals:
                        continue
                    dcs = res["dcs"]
                    bonuses = res["bonuses"]
                    min_roll_needed = max(max(dcs) - min(bonuses), 1)
                    self.tree_mon.insert(
                        parent, "end",
                        text="  Save: " + check_key,
                        values=("", "", "",
                                "Bonus {}-{} / DC {}-{} / "
                                "Needs {}+ / {}S-{}F".format(
                                    min(bonuses), max(bonuses),
                                    min(dcs), max(dcs), min_roll_needed,
                                    res["success"], res["fail"])))

                dealt = ex["dmg_dealt_types"]
                for dtype, amt in sorted(dealt.items()):
                    self.tree_mon.insert(
                        parent, "end",
                        text="  Damage Type: " + dtype,
                        values=("", "", "", "Dealt {}".format(amt)))

            if d.get("kills"):
                victims = ", ".join(d.get("kill_victims", []))
                self.tree_mon.insert(
                    parent, "end",
                    text="  Kills",
                    values=("", "", "",
                            "{} kill(s): {}".format(d["kills"], victims)))

            if ex:
                spells = ex.get("spells_cast")
                if spells:
                    for spell, count in sorted(spells.items()):
                        self.tree_mon.insert(
                            parent, "end",
                            text="  Cast: " + spell,
                            values=("", "", "", "{}x".format(count)))

                atk_h = ex.get("atk_hits", 0)
                atk_m = ex.get("atk_misses", 0)
                def_h = ex.get("def_hits", 0)
                def_m = ex.get("def_misses", 0)
                if atk_h or atk_m or def_h or def_m:
                    self.tree_mon.insert(
                        parent, "end",
                        text="  Attack Record",
                        values=("", "", "",
                                "Swings {}H-{}M / Taken {}H-{}M".format(
                                    atk_h, atk_m, def_h, def_m)))



    def _fill_saves_tab(self, extra_stats, characters):
        self._clear_tree(self.tree_saves)
        for name, d in sorted(extra_stats.items()):
            if name in IGNORED_NAMES:
                continue
            category = "Character" if name in characters else "Monster"
            for check_key, res in sorted(d["checks"].items()):
                totals = res["totals"]
                if not totals:
                    continue
                dcs = res["dcs"]
                bonuses = res["bonuses"]
                min_roll_needed = max(max(dcs) - min(bonuses), 1)
                self.tree_saves.insert(
                    "", "end",
                    values=(
                        name, category, check_key, len(totals),
                        "{} - {}".format(min(bonuses), max(bonuses)),
                        "{} - {}".format(min(dcs), max(dcs)),
                        min_roll_needed,
                        "{}/{}".format(res["success"], res["fail"])))

    def _fill_mitigation_tab(self, extra_stats, characters):
        self._clear_tree(self.tree_mit)
        for name, d in sorted(extra_stats.items()):
            if name in IGNORED_NAMES:
                continue
            has_data = (d["dr_absorbed"] or d["immunity_absorbed"]
                        or d["concealed_against"] or d["threat_rolls"])
            if not has_data:
                continue
            category = "Character" if name in characters else "Monster"
            avg_pct = (sum(d["concealment_pcts"]) /
                       len(d["concealment_pcts"])) \
                if d["concealment_pcts"] else 0
            tr = d["threat_rolls"]
            self.tree_mit.insert(
                "", "end",
                values=(
                    name, category, d["dr_absorbed"],
                    d["immunity_absorbed"], d["concealed_against"],
                    round(avg_pct, 1), len(tr),
                    max(tr) if tr else 0,
                    round(sum(tr) / len(tr), 2) if tr else 0))

    def _fill_damage_types_tab(self, extra_stats, characters):
        self._clear_tree(self.tree_dmg_types)
        for name, d in sorted(extra_stats.items()):
            if name in IGNORED_NAMES:
                continue
            dealt = d["dmg_dealt_types"]
            taken = d["dmg_taken_types"]
            if not dealt and not taken:
                continue
            category = "Character" if name in characters else "Monster"
            all_types = sorted(set(dealt.keys()) | set(taken.keys()))
            for dtype in all_types:
                self.tree_dmg_types.insert(
                    "", "end",
                    values=(
                        name, category, dtype,
                        dealt.get(dtype, 0), taken.get(dtype, 0)))

    def _fill_leaderboard_tab(self, stats_pc):
        self._clear_tree(self.tree_leaderboard)
        ranked = sorted(
            stats_pc.items(),
            key=lambda kv: kv[1]["dmg_tot"], reverse=True)
        for rank, (name, d) in enumerate(ranked, start=1):
            avg = d["dmg_tot"] / d["hits_count"] if d["hits_count"] else 0
            self.tree_leaderboard.insert(
                "", "end",
                values=(
                    rank, name, d["dmg_tot"], d["kills"],
                    round(avg, 2)))


    # ---------- Events window ----------
    def open_events_window(self):
        if not self.extra_stats:
            messagebox.showwarning(
                "No data", "Run an analysis first.")
            return

        win = tk.Toplevel(self)
        win.title("Combat Events")
        win.geometry("900x560")

        top = ttk.Frame(win)
        top.pack(fill="x", padx=8, pady=6)

        ttk.Label(top, text="Name:").pack(side="left")
        names = sorted(
            n for n in self.extra_stats
            if n not in IGNORED_NAMES and self.extra_stats[n]["events"]
        )
        name_var = tk.StringVar(value=names[0] if names else "")
        name_box = ttk.Combobox(
            top, textvariable=name_var, values=names,
            state="readonly", width=30)
        name_box.pack(side="left", padx=6)

        filter_var = tk.StringVar(value="all")
        ttk.Radiobutton(
            top, text="All", variable=filter_var,
            value="all").pack(side="left", padx=4)
        ttk.Radiobutton(
            top, text="Hits", variable=filter_var,
            value="hit").pack(side="left", padx=4)
        ttk.Radiobutton(
            top, text="Misses", variable=filter_var,
            value="miss").pack(side="left", padx=4)

        list_frame = ttk.Frame(win)
        list_frame.pack(fill="both", expand=True, padx=8, pady=6)
        events_list = tk.Listbox(list_frame, font=("Consolas", 9))
        events_list.pack(side="left", fill="both", expand=True)
        vsb = ttk.Scrollbar(
            list_frame, orient="vertical", command=events_list.yview)
        events_list.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")

        def refresh(*_args):
            events_list.delete(0, "end")
            name = name_var.get()
            if not name:
                return
            kind = filter_var.get()
            for ev_kind, raw in self.extra_stats[name]["events"]:
                if kind != "all" and ev_kind != kind:
                    continue
                events_list.insert("end", raw)

        name_box.bind("<<ComboboxSelected>>", refresh)
        filter_var.trace_add("write", refresh)
        refresh()

    # ---------- Excel export ----------
    def export_excel(self):
        if not self.stats_pc and not self.stats_m:
            messagebox.showwarning(
                "No data", "Run an analysis first.")
            return

        path = filedialog.asksaveasfilename(
            title="Save analysis as",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")])
        if not path:
            return

        wb = Workbook()
        self._write_character_sheet(wb.active)
        wb.active.title = "Character Stats"
        self._write_monster_sheet(wb.create_sheet("Monster Stats"))
        self._write_saves_sheet(wb.create_sheet("Saving Throws"))
        self._write_leaderboard_sheet(wb.create_sheet("Leaderboard"))
        wb.save(path)

        messagebox.showinfo("Exported", "Analysis saved to:\n" + path)

    def _write_character_sheet(self, ws):
        ws.append([
            "Name", "Kills", "Total Dmg", "Avg/Hit",
            "Atk Hits", "Atk Misses", "Def Hits", "Def Misses"])
        for name in self.characters:
            d = self.stats_pc.get(name, {})
            ex = self.extra_stats.get(name, {})
            avg = (d.get("dmg_tot", 0) / d["hits_count"]
                   if d.get("hits_count") else 0)
            ws.append([
                name, d.get("kills", 0), d.get("dmg_tot", 0),
                round(avg, 2), ex.get("atk_hits", 0),
                ex.get("atk_misses", 0), ex.get("def_hits", 0),
                ex.get("def_misses", 0)])

    def _write_monster_sheet(self, ws):
        ws.append([
            "Name", "Max AB", "Avg Dmg", "Kills",
            "Atk Hits", "Atk Misses", "Def Hits", "Def Misses"])
        for name, d in sorted(self.stats_m.items()):
            if name in self.characters or name in IGNORED_NAMES:
                continue
            ex = self.extra_stats.get(name, {})
            avg_m = (d["dmg_val"] / d["dmg_count"]
                     if d.get("dmg_count") else 0)
            ws.append([
                name, max(d["ab"]) if d.get("ab") else 0,
                round(avg_m, 2), d.get("kills", 0),
                ex.get("atk_hits", 0), ex.get("atk_misses", 0),
                ex.get("def_hits", 0), ex.get("def_misses", 0)])

    def _write_saves_sheet(self, ws):
        ws.append([
            "Name", "Category", "Check", "Count",
            "Bonus Range", "DC Range", "Min Roll Needed",
            "Success", "Fail"])
        for name, d in sorted(self.extra_stats.items()):
            if name in IGNORED_NAMES:
                continue
            category = "Character" if name in self.characters \
                else "Monster"
            for check_key, res in sorted(d["checks"].items()):
                totals = res["totals"]
                if not totals:
                    continue
                dcs, bonuses = res["dcs"], res["bonuses"]
                needed = max(max(dcs) - min(bonuses), 1)
                ws.append([
                    name, category, check_key, len(totals),
                    "{}-{}".format(min(bonuses), max(bonuses)),
                    "{}-{}".format(min(dcs), max(dcs)), needed,
                    res["success"], res["fail"]])

    def _write_leaderboard_sheet(self, ws):
        ws.append(["Rank", "Character", "Total Damage", "Kills", "Avg/Hit"])
        ranked = sorted(
            self.stats_pc.items(),
            key=lambda kv: kv[1]["dmg_tot"], reverse=True)
        for rank, (name, d) in enumerate(ranked, start=1):
            avg = d["dmg_tot"] / d["hits_count"] if d["hits_count"] else 0
            ws.append([rank, name, d["dmg_tot"], d["kills"], round(avg, 2)])


if __name__ == "__main__":
    _fix_windows_dpi_scaling()
    app = NWNAnalyzerApp()
    app.mainloop()
